import tkinter as tk
from tkinter import messagebox
import subprocess
import threading
import re
from datetime import datetime
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from license_validator import validar_licencia

if not validar_licencia():
    import sys
    sys.exit(1)

RUTA_LOGS = os.path.join(os.path.expanduser("~"), "Documents", "PING_LOGS")
RUTA_GRAFICOS = os.path.join(os.path.expanduser("~"), "Documents", "PING_GRAFICOS")
os.makedirs(RUTA_LOGS, exist_ok=True)
os.makedirs(RUTA_GRAFICOS, exist_ok=True)


class PingApp:
    def __init__(self, master, log_area, promedios_area, results):
        self.nombres_equipos = {
            "10.72.14.81": "EX 002",
            "10.72.14.82": "EX 003",
            "10.72.14.83": "EX 004",
            "10.72.14.84": "EX 005",
            "10.72.14.85": "EX 006",
            "10.72.14.86": "EX 066",
            "10.72.14.87": "EX 070",
            "10.72.14.88": "EX 071"
        }

        self.master = master
        self.log_area = log_area
        self.promedios_area = promedios_area
        self.results = results
        self.proceso_ping = None
        self.hora_inicio = None
        self.correctos = 0
        self.medios = 0
        self.medios_con_perdida = 0
        self.perdidos = 0
        self.ip = ""

        self.label = tk.Label(master, text="Ingresa la IP a hacer ping:")
        self.label.pack(pady=10)

        self.entry_ip = tk.Entry(master, width=30)
        self.entry_ip.insert(0, "10.72.14.")
        self.entry_ip.pack(pady=10)

        self.label_nombre_equipo = tk.Label(master, text="Equipo: Desconocido")
        self.label_nombre_equipo.pack(pady=10)

        self.btn_iniciar = tk.Button(master, text="Iniciar Ping", command=self.iniciar_ping)
        self.btn_iniciar.pack(pady=10)

        self.btn_detener = tk.Button(master, text="Detener Ping", command=self.detener_ping, state=tk.DISABLED)
        self.btn_detener.pack(pady=10)

        self.output_text = tk.Text(master, height=7, width=70)
        self.output_text.pack(pady=10)

        # --- Gráfico en tiempo real ---
        self.tiempos = []
        self.max_points = 100

        self.frame_grafico = tk.Frame(master)
        self.frame_grafico.pack(pady=10)

        self.fig, self.ax = plt.subplots(figsize=(8, 4))
        self.fig.patch.set_facecolor('#f5f5f5')
        self.ax.set_facecolor('#ffffff')

        self.linea, = self.ax.plot([], [], color='#1976D2', linewidth=2, alpha=0.8, label="Tendencia")
        self.scatter = None

        self.ax.set_title("Monitoreo de Latencia en Tiempo Real", fontsize=14, fontweight='bold', color='#212121', pad=15)
        self.ax.set_xlabel("Tiempo transcurrido (s)", fontsize=11, fontweight='600', color='#424242')
        self.ax.set_ylabel("Latencia (ms)", fontsize=11, fontweight='600', color='#424242')
        self.ax.grid(True, linestyle='--', alpha=0.3, color='#BDBDBD')
        self.ax.spines['top'].set_visible(False)
        self.ax.spines['right'].set_visible(False)
        self.ax.spines['left'].set_color('#9E9E9E')
        self.ax.spines['bottom'].set_color('#9E9E9E')

        self.canvas_grafico = FigureCanvasTkAgg(self.fig, master=self.frame_grafico)
        self.canvas_grafico.draw()
        self.canvas_grafico.get_tk_widget().pack()

        self.master.protocol("WM_DELETE_WINDOW", self.on_closing)

    def actualizar_grafico(self, nuevo_tiempo):
        # --- CONFIGURACIÓN DEL GRÁFICO CON TIEMPO EN SEGUNDOS ---
        limite_superior = 80
        # Guardar el tiempo real para la clasificación de colores
        tiempo_real = nuevo_tiempo
        # Limitar el valor mostrado en el gráfico para que no exceda el eje Y
        tiempo_mostrado = min(nuevo_tiempo, limite_superior)

        # Guardar tiempo actual (para eje X)
        if not hasattr(self, 'start_time'):
            from time import time
            self.start_time = time()
            self.tiempos_segundos = []
            self.tiempos_reales = []  # Lista para almacenar valores reales

        from time import time
        tiempo_actual = time() - self.start_time
        self.tiempos_segundos.append(tiempo_actual)
        if len(self.tiempos_segundos) > self.max_points:
            self.tiempos_segundos.pop(0)
            self.tiempos.pop(0)
            self.tiempos_reales.pop(0)

        # Guardar ambos valores: mostrado y real
        self.tiempos.append(tiempo_mostrado)
        self.tiempos_reales.append(tiempo_real)

        # Actualizar línea y puntos
        self.linea.set_xdata(self.tiempos_segundos)
        self.linea.set_ydata(self.tiempos)

        # Colorear puntos según valor REAL (no el limitado) con 4 rangos de colores
        def obtener_color(t):
            if t < 40:
                return '#43A047'  # Verde - Correctos
            elif t < 50:
                return '#FDD835'  # Amarillo - Medios
            elif t < 120:
                return '#FF9800'  # Naranja - Medios con pérdida
            else:
                return "#FF0400"  # Rojo - Perdidos

        # Usar tiempos_reales para determinar el color
        colores = [obtener_color(t) for t in self.tiempos_reales]
        tamaños = [60 if t < 40 else 70 if t < 50 else 75 if t < 120 else 80 for t in self.tiempos_reales]

        if self.scatter:
            self.scatter.remove()
        self.scatter = self.ax.scatter(self.tiempos_segundos, self.tiempos, c=colores,
                                       s=tamaños, alpha=0.7, edgecolors='white', linewidths=1.5)

        # Ejes dinámicos
        if len(self.tiempos_segundos) > 1:
            self.ax.set_xlim(self.tiempos_segundos[0], self.tiempos_segundos[-1])

        # Fijar el límite del eje Y a 80ms
        self.ax.set_ylim(0, limite_superior)
        self.ax.set_yticks(range(0, limite_superior + 10, 10))

        self.canvas_grafico.draw_idle()


    def on_closing(self):
        if self.proceso_ping is not None:
            self.proceso_ping.terminate()
            self.output_text.insert(tk.END, "\nPing detenido al cerrar la ventana.\n")
        self.master.destroy()

    def iniciar_ping(self):
        if self.proceso_ping is not None:
            messagebox.showerror("Error", "Ya hay un ping en curso.")
            return

        self.ip = self.entry_ip.get()
        if not self.ip:
            messagebox.showerror("Error", "Por favor ingrese una IP válida.")
            return

        nombre_equipo = self.nombres_equipos.get(self.ip, "Desconocido")
        self.label_nombre_equipo.config(text=f"Equipo: {nombre_equipo}")

        self.correctos = 0
        self.medios = 0
        self.medios_con_perdida = 0
        self.perdidos = 0
        self.hora_inicio = datetime.now()

        # Reiniciar el gráfico completamente
        self.tiempos.clear()
        self.tiempos_segundos = []
        if hasattr(self, 'tiempos_reales'):
            self.tiempos_reales.clear()
        if hasattr(self, 'start_time'):
            delattr(self, 'start_time')

        self.linea.set_xdata([])
        self.linea.set_ydata([])
        if self.scatter:
            self.scatter.remove()
            self.scatter = None
        self.ax.set_xlim(0, 10)
        self.ax.set_ylim(0, 80)
        self.canvas_grafico.draw_idle()

        self.btn_detener.config(state=tk.NORMAL)

        self.output_text.delete(1.0, tk.END)
        self.output_text.insert(tk.END, f"Iniciando ping a {self.ip} ({nombre_equipo})...\n")

        def ping():
            creation_flags = 0
            if os.name == 'nt':
                creation_flags = subprocess.CREATE_NO_WINDOW

            self.proceso_ping = subprocess.Popen(
                ["ping", self.ip, "-t"] if os.name == 'nt' else ["ping", self.ip],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                creationflags=creation_flags
            )

            try:
                while True:
                    line = self.proceso_ping.stdout.readline()
                    if line == '':
                        break

                    self.output_text.insert(tk.END, line)
                    self.output_text.yview(tk.END)

                    match = re.search(r"(time|tiempo)[=<](\d+)ms", line, re.IGNORECASE)
                    if match:
                        tiempo = int(match.group(2))
                        # Clasificación en 4 rangos
                        if tiempo < 40:
                            self.correctos += 1
                        elif tiempo < 50:
                            self.medios += 1
                        elif tiempo < 120:
                            self.medios_con_perdida += 1
                        else:
                            self.perdidos += 1

                        # 🚀 actualizar gráfico
                        self.master.after(0, self.actualizar_grafico, tiempo)

            finally:
                self.proceso_ping = None

        threading.Thread(target=ping, daemon=True).start()

    def detener_ping(self):
        if self.proceso_ping is not None:
            self.proceso_ping.terminate()
            hora_fin = datetime.now()
            duracion = hora_fin - self.hora_inicio

            horas, resto = divmod(duracion.total_seconds(), 3600)
            minutos, segundos = divmod(resto, 60)
            tiempo = f"{int(horas)}h, {int(minutos)}m, {int(segundos)}s"

            nombre_equipo = self.nombres_equipos.get(self.ip, "Desconocido")

            self.output_text.insert(tk.END, "\nProceso de ping detenido por el usuario.\n")

            total_paquetes = self.correctos + self.medios + self.medios_con_perdida + self.perdidos
            if total_paquetes > 0:
                porcentaje_correctos = (self.correctos / total_paquetes) * 100
                porcentaje_medios = (self.medios / total_paquetes) * 100
                porcentaje_medios_perdida = (self.medios_con_perdida / total_paquetes) * 100
                porcentaje_perdidos = (self.perdidos / total_paquetes) * 100
            else:
                porcentaje_correctos = 0
                porcentaje_medios = 0
                porcentaje_medios_perdida = 0
                porcentaje_perdidos = 0

            self.output_text.insert(tk.END, f"\nDesde: {self.hora_inicio.strftime('%Y-%m-%d %H:%M:%S')}\n")
            self.output_text.insert(tk.END, f"Hasta: {hora_fin.strftime('%Y-%m-%d %H:%M:%S')}\n")
            self.output_text.insert(tk.END, f"Duración total: {tiempo}\n")
            self.output_text.insert(tk.END, f"% Correctos (< 40ms): {porcentaje_correctos:.2f}%\n")
            self.output_text.insert(tk.END, f"% Medios (40-50ms): {porcentaje_medios:.2f}%\n")
            self.output_text.insert(tk.END, f"% Medios con Pérdida (50-120ms): {porcentaje_medios_perdida:.2f}%\n")
            self.output_text.insert(tk.END, f"% Perdidos (>= 120ms): {porcentaje_perdidos:.2f}%\n")

            self.log_area.insert(tk.END, f"--- Resultados de Ping para IP: {self.ip} ({nombre_equipo}) ---\n")
            self.log_area.insert(tk.END, f"Desde: {self.hora_inicio.strftime('%Y-%m-%d %H:%M:%S')}\n")
            self.log_area.insert(tk.END, f"Hasta: {hora_fin.strftime('%Y-%m-%d %H:%M:%S')}\n")
            self.log_area.insert(tk.END, f"Duración total: {tiempo}\n\n")
            self.log_area.insert(tk.END, f"% Correctos (< 40ms): {porcentaje_correctos:.2f}%\n")
            self.log_area.insert(tk.END, f"% Medios (40-50ms): {porcentaje_medios:.2f}%\n")
            self.log_area.insert(tk.END, f"% Medios con Pérdida (50-120ms): {porcentaje_medios_perdida:.2f}%\n")
            self.log_area.insert(tk.END, f"% Perdidos (>= 120ms): {porcentaje_perdidos:.2f}%\n\n")

            fecha_hora = datetime.now().strftime("%d%m%Y_%H%M%S")
            nombre_archivo = os.path.join(RUTA_LOGS, f"{nombre_equipo}_{fecha_hora}.txt")
            with open(nombre_archivo, "a", encoding="utf-8") as f:
                f.write(f"\n==============================\n")
                f.write(f"PING a {self.ip} - {nombre_equipo}\n")
                f.write(f"Desde: {self.hora_inicio.strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Hasta: {hora_fin.strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Duración total: {tiempo}\n")
                f.write("------------------------------\n")
                f.write(self.output_text.get("1.0", tk.END))
                f.write("\n==============================\n\n")

            self.results.append({
                'ip': self.ip,
                'nombre': nombre_equipo,
                'correctos': porcentaje_correctos,
                'medios': porcentaje_medios,
                'medios_con_perdida': porcentaje_medios_perdida,
                'perdidos': porcentaje_perdidos,
                'tiempos_grafico': self.tiempos_reales.copy() if hasattr(self, 'tiempos_reales') else self.tiempos.copy(),
                'tiempos_segundos_grafico': self.tiempos_segundos.copy() if hasattr(self, 'tiempos_segundos') else []
            })

            self.tiempos.clear()
            self.linea.set_xdata([])
            self.linea.set_ydata([])
            if self.scatter:
                self.scatter.remove()
                self.scatter = None
            self.ax.set_xlim(0, self.max_points)
            self.ax.set_ylim(0, 100)
            self.canvas_grafico.draw_idle()

            self.btn_detener.config(state=tk.DISABLED)
            self.proceso_ping = None
        else:
            self.output_text.insert(tk.END, "\nNo hay un ping en curso para detener.\n")

def limpiar_log():
    log_area.delete(1.0, tk.END)
    promedios_area.delete(1.0, tk.END)
    results.clear()


def calcular_promedio():
    log_text = log_area.get(1.0, tk.END)
    total_correctos = 0
    total_medios = 0
    total_medios_perdida = 0
    total_perdidos = 0
    conteos = 0

    # Reiniciar la lista results en base a lo que haya en log_area
    results.clear()

    for line in log_text.splitlines():
        if "--- Resultados de Ping para IP:" in line:
            conteos += 1
            ip = re.search(r"IP: ([\d.]+)", line)
            nombre = re.search(r"\((.*?)\)", line)
            ip_value = ip.group(1) if ip else "Desconocida"
            nombre_value = nombre.group(1) if nombre else "Desconocido"

            # Crear una entrada temporal que luego será completada
            results.append({
                'ip': ip_value,
                'nombre': nombre_value,
                'correctos': 0,
                'medios': 0,
                'medios_con_perdida': 0,
                'perdidos': 0
            })

        if "% Correctos (< 40ms)" in line:
            match = re.search(r"% Correctos \(< 40ms\): (\d+.\d+)%", line)
            if match and results:
                val = float(match.group(1))
                total_correctos += val
                results[-1]['correctos'] = val

        if "% Medios (40-50ms)" in line:
            match = re.search(r"% Medios \(40-50ms\): (\d+.\d+)%", line)
            if match and results:
                val = float(match.group(1))
                total_medios += val
                results[-1]['medios'] = val

        if "% Medios con Pérdida (50-120ms)" in line:
            match = re.search(r"% Medios con Pérdida \(50-120ms\): (\d+.\d+)%", line)
            if match and results:
                val = float(match.group(1))
                total_medios_perdida += val
                results[-1]['medios_con_perdida'] = val

        if "% Perdidos (>= 120ms)" in line:
            match = re.search(r"% Perdidos \(>= 120ms\): (\d+.\d+)%", line)
            if match and results:
                val = float(match.group(1))
                total_perdidos += val
                results[-1]['perdidos'] = val

    if conteos > 0:
        promedio_correctos = total_correctos / conteos
        promedio_medios = total_medios / conteos
        promedio_medios_perdida = total_medios_perdida / conteos
        promedio_perdidos = total_perdidos / conteos

        # Actualizar área de promedios
        promedios_area.delete(1.0, tk.END)
        promedios_area.insert(tk.END, f"Promedio % Correctos (< 40ms): {promedio_correctos:.2f}%\n")
        promedios_area.insert(tk.END, f"Promedio % Medios (40-50ms): {promedio_medios:.2f}%\n")
        promedios_area.insert(tk.END, f"Promedio % Medios con Pérdida (50-120ms): {promedio_medios_perdida:.2f}%\n")
        promedios_area.insert(tk.END, f"Promedio % Perdidos (>= 120ms): {promedio_perdidos:.2f}%\n")
    else:
        promedios_area.delete(1.0, tk.END)
        promedios_area.insert(tk.END, "No hay datos suficientes.\n")



def generar_grafico():
    if not results:
        messagebox.showerror("Error", "No hay resultados para mostrar.")
        return

    ips = [result['ip'] for result in results]
    nombres = [result['nombre'] for result in results]
    correctos = [result['correctos'] for result in results]
    medios = [result['medios'] for result in results]
    medios_con_perdida = [result['medios_con_perdida'] for result in results]
    perdidos = [result['perdidos'] for result in results]

    promedio_correctos = sum(correctos) / len(correctos) if correctos else 0
    promedio_medios = sum(medios) / len(medios) if medios else 0
    promedio_medios_perdida = sum(medios_con_perdida) / len(medios_con_perdida) if medios_con_perdida else 0
    promedio_perdidos = sum(perdidos) / len(perdidos) if perdidos else 0

    new_window = tk.Toplevel(ventana_principal)
    new_window.title("Gráficos de Resultados de Ping")

    # Frame para botones en la parte superior
    button_frame = tk.Frame(new_window)
    button_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)

    # Lista para guardar las figuras
    figuras_guardadas = []

    canvas_frame = tk.Frame(new_window)
    canvas_frame.pack(fill=tk.BOTH, expand=True)

    canvas = tk.Canvas(canvas_frame)
    scrollbar = tk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=scrollbar.set)

    scrollbar.pack(side="right", fill="y")
    canvas.pack(side="left", fill=tk.BOTH, expand=True)

    graphics_frame = tk.Frame(canvas)
    canvas.create_window((0, 0), window=graphics_frame, anchor="nw")

    # Primero: Gráficos individuales de cada pineado
    for result in results:
        if result.get('tiempos_grafico') and result.get('tiempos_segundos_grafico'):
            fig_individual, ax_individual = plt.subplots(figsize=(10, 4))
            fig_individual.patch.set_facecolor('#f5f5f5')
            ax_individual.set_facecolor('#ffffff')

            tiempos_reales = result['tiempos_grafico']  # Estos son los valores reales
            tiempos_segundos = result['tiempos_segundos_grafico']

            # Limitar los valores para mostrar en el gráfico (máximo 80ms)
            tiempos_limitados = [min(t, 80) for t in tiempos_reales]

            # Línea de tendencia (con valores limitados)
            ax_individual.plot(tiempos_segundos, tiempos_limitados, color='#1976D2', linewidth=2, alpha=0.8, label="Tendencia")

            # Puntos coloreados con 4 rangos (usar valores REALES para colores)
            def obtener_color(t):
                if t < 40:
                    return '#43A047'  # Verde - Correctos
                elif t < 50:
                    return '#FDD835'  # Amarillo - Medios
                elif t < 120:
                    return '#FF9800'  # Naranja - Medios con pérdida
                else:
                    return "#FC0400"  # Rojo - Perdidos

            colores = [obtener_color(t) for t in tiempos_reales]
            tamaños = [60 if t < 40 else 70 if t < 50 else 75 if t < 120 else 80 for t in tiempos_reales]
            ax_individual.scatter(tiempos_segundos, tiempos_limitados, c=colores,
                                 s=tamaños, alpha=0.7, edgecolors='white', linewidths=1.5)

            ax_individual.set_title(f"Historial de Latencia - {result['nombre']} ({result['ip']})",
                                   fontsize=12, fontweight='bold', color='#212121', pad=12)
            ax_individual.set_xlabel("Tiempo transcurrido (s)", fontsize=10, fontweight='600', color='#424242')
            ax_individual.set_ylabel("Latencia (ms)", fontsize=10, fontweight='600', color='#424242')
            ax_individual.grid(True, linestyle='--', alpha=0.3, color='#BDBDBD')
            ax_individual.spines['top'].set_visible(False)
            ax_individual.spines['right'].set_visible(False)
            ax_individual.spines['left'].set_color('#9E9E9E')
            ax_individual.spines['bottom'].set_color('#9E9E9E')

            if tiempos_segundos:
                ax_individual.set_xlim(tiempos_segundos[0], tiempos_segundos[-1])
            # Fijar el límite del eje Y a 80ms
            ax_individual.set_ylim(0, 80)

            plt.tight_layout()

            # Guardar figura para exportación posterior
            figuras_guardadas.append({
                'figura': fig_individual,
                'nombre': f"Historial_{result['nombre'].replace(' ', '_')}",
                'tipo': 'individual'
            })

            canvas_individual = FigureCanvasTkAgg(fig_individual, master=graphics_frame)
            canvas_individual.draw()
            canvas_individual.get_tk_widget().pack(pady=10)

    # Segundo: Gráfico de barras comparativo
    fig1, ax1 = plt.subplots(figsize=(12, 6))
    fig1.patch.set_facecolor('#f8f9fa')
    ax1.set_facecolor('#ffffff')

    bar_width = 0.2
    index = range(len(ips))

    # Colores profesionales para los 4 rangos
    color_correctos = '#2E7D32'  # Verde oscuro
    color_medios = '#F9A825'  # Amarillo oscuro
    color_medios_perdida = '#EF6C00'  # Naranja oscuro
    color_perdidos = "#FF0000"  # Rojo oscuro

    # Crear 4 barras con diseño moderno
    bar1 = ax1.bar(index, correctos, bar_width, label='Correctos (< 40ms)',
                   color=color_correctos, edgecolor='#1B5E20', linewidth=1.5, alpha=0.85)
    bar2 = ax1.bar([i + bar_width for i in index], medios, bar_width,
                   label='Medios (40-50ms)', color=color_medios,
                   edgecolor='#F57F17', linewidth=1.5, alpha=0.85)
    bar3 = ax1.bar([i + bar_width * 2 for i in index], medios_con_perdida, bar_width,
                   label='Medios con Pérdida (50-120ms)', color=color_medios_perdida,
                   edgecolor='#E65100', linewidth=1.5, alpha=0.85)
    bar4 = ax1.bar([i + bar_width * 3 for i in index], perdidos, bar_width,
                   label='Perdidos (≥ 120ms)', color=color_perdidos,
                   edgecolor='#B71C1C', linewidth=1.5, alpha=0.85)

    def add_labels(bars, data):
        for bar, value in zip(bars, data):
            height = bar.get_height()
            if height > 0:  # Solo mostrar etiqueta si hay valor
                ax1.text(bar.get_x() + bar.get_width() / 2, height + 1.5,
                        f'{value:.1f}%', ha='center', va='bottom',
                        fontsize=8, fontweight='bold', color='#263238')

    add_labels(bar1, correctos)
    add_labels(bar2, medios)
    add_labels(bar3, medios_con_perdida)
    add_labels(bar4, perdidos)

    ax1.set_xlabel('Equipos de Excavación', fontsize=11, fontweight='bold', color='#263238', labelpad=8)
    ax1.set_ylabel('Porcentaje de Paquetes (%)', fontsize=11, fontweight='bold', color='#263238', labelpad=8)
    ax1.set_title('Análisis de Rendimiento de Red por Equipo', fontsize=13, fontweight='bold',
                  color='#1A237E', pad=15)
    ax1.set_xticks([i + bar_width * 1.5 for i in index])
    ax1.set_xticklabels(nombres, rotation=35, ha='right', fontsize=9, fontweight='600')
    ax1.legend(title="Categorías de Latencia", title_fontsize='10', fontsize=9,
              loc='upper right', framealpha=0.95, edgecolor='#BDBDBD')

    ax1.grid(True, linestyle='--', alpha=0.25, color='#9E9E9E', axis='y')
    ax1.set_ylim(0, 105)

    # Mejorar los bordes del gráfico
    ax1.spines['top'].set_visible(False)
    ax1.spines['right'].set_visible(False)
    ax1.spines['left'].set_color('#757575')
    ax1.spines['bottom'].set_color('#757575')
    ax1.spines['left'].set_linewidth(1.5)
    ax1.spines['bottom'].set_linewidth(1.5)

    ax1.tick_params(axis='both', which='major', labelsize=9, colors='#424242')

    plt.tight_layout()

    # Guardar figura comparativa
    figuras_guardadas.append({
        'figura': fig1,
        'nombre': 'Comparativo_Equipos',
        'tipo': 'comparativo'
    })

    canvas1 = FigureCanvasTkAgg(fig1, master=graphics_frame)
    canvas1.draw()
    canvas1.get_tk_widget().pack(pady=10)

    # Tercero: Gráfico de promedio general
    fig2, ax2 = plt.subplots(figsize=(12, 6))
    fig2.patch.set_facecolor('#f8f9fa')
    ax2.set_facecolor('#ffffff')

    bar_width_avg = 0.4
    x_pos = [0, 0.5, 1.0, 1.5]

    # Barras profesionales para el promedio con 4 categorías
    bar1_avg = ax2.bar(x_pos[0], promedio_correctos, bar_width_avg,
                       label='Correctos (< 40ms)', color=color_correctos,
                       edgecolor='#1B5E20', linewidth=2, alpha=0.85)
    bar2_avg = ax2.bar(x_pos[1], promedio_medios, bar_width_avg,
                       label='Medios (40-50ms)', color=color_medios,
                       edgecolor='#F57F17', linewidth=2, alpha=0.85)
    bar3_avg = ax2.bar(x_pos[2], promedio_medios_perdida, bar_width_avg,
                       label='Medios con Pérdida (50-120ms)', color=color_medios_perdida,
                       edgecolor='#E65100', linewidth=2, alpha=0.85)
    bar4_avg = ax2.bar(x_pos[3], promedio_perdidos, bar_width_avg,
                       label='Perdidos (≥ 120ms)', color=color_perdidos,
                       edgecolor="#FF0000", linewidth=2, alpha=0.85)

    # Etiquetas grandes y destacadas
    ax2.text(x_pos[0], promedio_correctos + 2,
            f'{promedio_correctos:.2f}%', ha='center', va='bottom',
            fontsize=13, fontweight='bold', color='#1B5E20')
    ax2.text(x_pos[1], promedio_medios + 2,
            f'{promedio_medios:.2f}%', ha='center', va='bottom',
            fontsize=13, fontweight='bold', color='#F57F17')
    ax2.text(x_pos[2], promedio_medios_perdida + 2,
            f'{promedio_medios_perdida:.2f}%', ha='center', va='bottom',
            fontsize=13, fontweight='bold', color='#E65100')
    ax2.text(x_pos[3], promedio_perdidos + 2,
            f'{promedio_perdidos:.2f}%', ha='center', va='bottom',
            fontsize=13, fontweight='bold', color='#B71C1C')

    # Añadir información adicional
    total_tests = len(ips)
    ax2.text(0.75, -12, f'Basado en {total_tests} equipo(s) monitoreado(s)',
            ha='center', va='top', fontsize=9, style='italic', color='#546E7A')

    ax2.set_ylabel('Porcentaje de Paquetes (%)', fontsize=11, fontweight='bold', color='#263238', labelpad=8)
    ax2.set_title('Rendimiento Promedio de la Red Mesh', fontsize=13, fontweight='bold',
                  color='#1A237E', pad=15)
    ax2.set_xticks(x_pos)
    ax2.set_xticklabels(['Óptimo', 'Bueno', 'Regular', 'Malo'],
                        fontsize=10, fontweight='600', color='#37474F')
    ax2.legend(title="Métricas de Red", title_fontsize='10', fontsize=9,
              loc='upper center', framealpha=0.95, edgecolor='#BDBDBD',
              bbox_to_anchor=(0.5, -0.12), ncol=4)

    ax2.grid(True, linestyle='--', alpha=0.25, color='#9E9E9E', axis='y')
    ax2.set_ylim(0, 110)
    ax2.set_xlim(-0.3, 2.1)

    # Mejorar los bordes del gráfico
    ax2.spines['top'].set_visible(False)
    ax2.spines['right'].set_visible(False)
    ax2.spines['left'].set_color('#757575')
    ax2.spines['bottom'].set_color('#757575')
    ax2.spines['left'].set_linewidth(1.5)
    ax2.spines['bottom'].set_linewidth(1.5)

    ax2.tick_params(axis='both', which='major', labelsize=9, colors='#424242')

    plt.tight_layout()

    # Guardar figura de promedio
    figuras_guardadas.append({
        'figura': fig2,
        'nombre': 'Promedio_General',
        'tipo': 'promedio'
    })

    canvas2 = FigureCanvasTkAgg(fig2, master=graphics_frame)
    canvas2.draw()
    canvas2.get_tk_widget().pack(pady=10)

    graphics_frame.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))

    # Función para exportar a Excel
    def exportar_excel():
        fecha_hora = datetime.now().strftime("%d%m%Y_%H%M%S")
        nombre_archivo = os.path.join(RUTA_GRAFICOS, f"Cuadro_Resumen_{fecha_hora}.xlsx")

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "CUADRO RESUMEN"

        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        promedio_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        promedio_font = Font(bold=True, size=11)
        border_style = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Título principal
        ws.merge_cells('A1:J1')
        ws['A1'] = "CUADRO RESUMEN: CARGUÍO"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_alignment
        ws['A1'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Encabezados
        headers = ["EQUIPO", "IP", "Fecha", "Hora de Inicio", "Hora final", "Duración",
                   "< 40ms (%)", "40-50ms (%)", "50-120ms (%)", "≥ 120ms (%)"]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border_style

        # Datos de cada equipo
        log_text = log_area.get(1.0, tk.END)
        current_row = 3

        # Parsear datos del log_area para obtener información completa
        equipos_data = []
        current_equipo = {}

        for line in log_text.splitlines():
            if "--- Resultados de Ping para IP:" in line:
                if current_equipo:
                    equipos_data.append(current_equipo)

                ip_match = re.search(r"IP: ([\d.]+)", line)
                nombre_match = re.search(r"\((.*?)\)", line)
                current_equipo = {
                    'ip': ip_match.group(1) if ip_match else "N/A",
                    'nombre': nombre_match.group(1) if nombre_match else "N/A",
                    'fecha': '',
                    'hora_inicio': '',
                    'hora_fin': '',
                    'duracion': '',
                    'correctos': 0,
                    'medios': 0,
                    'medios_perdida': 0,
                    'perdidos': 0
                }

            if current_equipo:
                if "Desde:" in line:
                    fecha_hora_match = re.search(r"Desde: (\d{4}-\d{2}-\d{2}) (\d{2}:\d{2}:\d{2})", line)
                    if fecha_hora_match:
                        current_equipo['fecha'] = fecha_hora_match.group(1)
                        current_equipo['hora_inicio'] = fecha_hora_match.group(2)

                if "Hasta:" in line:
                    hora_match = re.search(r"Hasta: \d{4}-\d{2}-\d{2} (\d{2}:\d{2}:\d{2})", line)
                    if hora_match:
                        current_equipo['hora_fin'] = hora_match.group(1)

                if "Duración total:" in line:
                    duracion_match = re.search(r"Duración total: (.+)", line)
                    if duracion_match:
                        current_equipo['duracion'] = duracion_match.group(1)

                if "% Correctos (< 40ms)" in line:
                    match = re.search(r"% Correctos \(< 40ms\): (\d+.\d+)%", line)
                    if match:
                        current_equipo['correctos'] = float(match.group(1))

                if "% Medios (40-50ms)" in line:
                    match = re.search(r"% Medios \(40-50ms\): (\d+.\d+)%", line)
                    if match:
                        current_equipo['medios'] = float(match.group(1))

                if "% Medios con Pérdida (50-120ms)" in line:
                    match = re.search(r"% Medios con Pérdida \(50-120ms\): (\d+.\d+)%", line)
                    if match:
                        current_equipo['medios_perdida'] = float(match.group(1))

                if "% Perdidos (>= 120ms)" in line:
                    match = re.search(r"% Perdidos \(>= 120ms\): (\d+.\d+)%", line)
                    if match:
                        current_equipo['perdidos'] = float(match.group(1))

        if current_equipo:
            equipos_data.append(current_equipo)

        # Escribir datos de equipos
        for equipo in equipos_data:
            ws.cell(row=current_row, column=1, value=equipo['nombre']).border = border_style
            ws.cell(row=current_row, column=2, value=equipo['ip']).border = border_style
            ws.cell(row=current_row, column=3, value=equipo['fecha']).border = border_style
            ws.cell(row=current_row, column=4, value=equipo['hora_inicio']).border = border_style
            ws.cell(row=current_row, column=5, value=equipo['hora_fin']).border = border_style
            ws.cell(row=current_row, column=6, value=equipo['duracion']).border = border_style

            # Porcentajes con formato
            for col, val in enumerate([equipo['correctos'], equipo['medios'],
                                       equipo['medios_perdida'], equipo['perdidos']], start=7):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.number_format = '0.00'
                cell.alignment = center_alignment
                cell.border = border_style

            # Centrar columnas de texto
            for col in range(1, 7):
                ws.cell(row=current_row, column=col).alignment = center_alignment

            current_row += 1

        # Fila de PROMEDIO
        ws.cell(row=current_row, column=1, value="PROMEDIO").font = promedio_font
        ws.cell(row=current_row, column=1).fill = promedio_fill
        ws.cell(row=current_row, column=1).alignment = center_alignment
        ws.cell(row=current_row, column=1).border = border_style

        # Combinar celdas para PROMEDIO
        ws.merge_cells(f'A{current_row}:F{current_row}')

        # Calcular y escribir promedios
        if equipos_data:
            prom_correctos = sum(e['correctos'] for e in equipos_data) / len(equipos_data)
            prom_medios = sum(e['medios'] for e in equipos_data) / len(equipos_data)
            prom_medios_perdida = sum(e['medios_perdida'] for e in equipos_data) / len(equipos_data)
            prom_perdidos = sum(e['perdidos'] for e in equipos_data) / len(equipos_data)

            for col, val in enumerate([prom_correctos, prom_medios, prom_medios_perdida, prom_perdidos], start=7):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.number_format = '0.00'
                cell.font = promedio_font
                cell.fill = promedio_fill
                cell.alignment = center_alignment
                cell.border = border_style

        # Ajustar anchos de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12

        # Guardar archivo
        wb.save(nombre_archivo)

        messagebox.showinfo(
            "Exportación Exitosa",
            f"Se exportó el cuadro resumen a:\n{nombre_archivo}"
        )

        # Abrir el archivo automáticamente
        if os.name == 'nt':  # Windows
            os.startfile(nombre_archivo)

    # Función para exportar todos los gráficos
    def exportar_graficos():
        fecha_hora = datetime.now().strftime("%d%m%Y_%H%M%S")
        carpeta_reporte = os.path.join(RUTA_GRAFICOS, f"Reporte_{fecha_hora}")
        os.makedirs(carpeta_reporte, exist_ok=True)

        contador_exportados = 0
        for fig_data in figuras_guardadas:
            nombre_archivo = f"{fig_data['nombre']}_{fecha_hora}.png"
            ruta_completa = os.path.join(carpeta_reporte, nombre_archivo)
            fig_data['figura'].savefig(ruta_completa, dpi=300, bbox_inches='tight', facecolor='white')
            contador_exportados += 1

        messagebox.showinfo(
            "Exportación Exitosa",
            f"Se exportaron {contador_exportados} gráfico(s) a:\n{carpeta_reporte}"
        )

        # Abrir la carpeta automáticamente
        import subprocess
        if os.name == 'nt':  # Windows
            os.startfile(carpeta_reporte)
        elif os.name == 'posix':  # macOS/Linux
            subprocess.Popen(['xdg-open', carpeta_reporte])

    # Botón para exportar a Excel
    btn_exportar_excel = tk.Button(
        button_frame,
        text="📑 Exportar Cuadro Resumen (Excel)",
        command=exportar_excel,
        bg='#2E7D32',
        fg='white',
        font=('Arial', 10, 'bold'),
        padx=15,
        pady=8
    )
    btn_exportar_excel.pack(side=tk.LEFT, padx=5)

    # Botón para exportar gráficos
    btn_exportar = tk.Button(
        button_frame,
        text="📊 Exportar Todos los Gráficos (PNG)",
        command=exportar_graficos,
        bg='#4CAF50',
        fg='white',
        font=('Arial', 10, 'bold'),
        padx=15,
        pady=8
    )
    btn_exportar.pack(side=tk.LEFT, padx=5)

    # Etiqueta informativa
    info_label = tk.Label(
        button_frame,
        text=f"Total: {len(figuras_guardadas)} gráfico(s)",
        font=('Arial', 9),
        fg='#546E7A'
    )
    info_label.pack(side=tk.LEFT, padx=10)

    new_window.protocol("WM_DELETE_WINDOW", lambda: new_window.destroy())


def generar_nueva_ventana():
    nueva_ventana = tk.Toplevel(ventana_principal)
    nueva_app = PingApp(nueva_ventana, log_area, promedios_area, results)


ventana_principal = tk.Tk()
ventana_principal.title("Gestión de Ventanas de Ping")

results = []

log_area = tk.Text(ventana_principal, height=10, width=70, state=tk.NORMAL, undo=True)
log_area.pack(pady=10)

promedios_area = tk.Text(ventana_principal, height=5, width=70)
promedios_area.pack(pady=10)

btn_generar_ventana = tk.Button(ventana_principal, text="Generar Nueva Ventana de Ping", command=generar_nueva_ventana)
btn_generar_ventana.pack(pady=10)

btn_calcular_promedio = tk.Button(ventana_principal, text="Calcular Promedio de Pings", command=calcular_promedio)
btn_calcular_promedio.pack(pady=10)

btn_limpiar = tk.Button(ventana_principal, text="Limpiar Log y Resultados", command=limpiar_log)
btn_limpiar.pack(pady=10)

btn_generar_grafico = tk.Button(ventana_principal, text="Generar Gráfico de Resultados", command=generar_grafico)
btn_generar_grafico.pack(pady=10)


def cerrar_aplicacion():
    # Detener cualquier ping en ejecución
    for widget in ventana_principal.winfo_children():
        if isinstance(widget, tk.Toplevel):  # ventanas hijas
            widget.destroy()
    os._exit(0)  # mata todos los procesos asociados (incluyendo pings)
    
ventana_principal.protocol("WM_DELETE_WINDOW", cerrar_aplicacion)

ventana_principal.mainloop()
